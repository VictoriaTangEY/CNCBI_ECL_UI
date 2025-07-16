import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import itertools
import json


def _deal_sheet(df, _parm):
    """
    Casts the data types of columns in a parameter table DataFrame `df`,
    based on the specified data types provided in the row named 'Parameter data type'.

    Parameters:
    - df (DataFrame): The input DataFrame containing variable parameter table.
    - _parm (str): Parameter table sheet name.

    Returns:
    - DataFrame: Processed DataFrame with columns' data types converted according to the 'Parameter data type' row.
    """
    col_idx = df.iloc[:, 0].values.tolist()
    try:
        start_idx = col_idx.index(1)
    except Exception as e:
        print(f"Issue observed: no variable count number is 1 in {_parm}")
        print(f"Error as {e}")

    variable_params = df.iloc[start_idx:, :]

    try:
        type_idx = col_idx.index('Parameter data type')
    except Exception as e:
        print(
            f"Issue observed in finding data type row indicator 'Parameter data type'in {_parm} ")
        print(f"Error as {e}")
    param_data_types = df.iloc[type_idx]

    # Date parsing function
    def try_parse_date(series):
        """Try to parse dates using common formats"""
        # Try the most common formats in your data
        for fmt in ["%d%b%Y:%H:%M:%S", "%d/%m/%Y", "%Y-%m-%d", "%Y%m%d"]:
            try:
                parsed = pd.to_datetime(series, format=fmt, errors='coerce')
                # If at least some values were parsed, use this format
                if parsed.notna().sum() > 0:
                    return parsed
            except Exception:
                continue
        # Fallback: let pandas try to infer
        return pd.to_datetime(series, errors='coerce')

    for col in variable_params.columns:
        data_type = param_data_types[col]
        try:
            if data_type == 'Parameter data type':
                # Handle header row - convert to string
                variable_params[col] = variable_params[col].fillna(
                    '').astype(str)
            elif data_type == 'integer' or data_type == 'int' or data_type == int:
                # Use pandas nullable Int64 type to allow NA
                variable_params[col] = pd.to_numeric(
                    variable_params[col], errors='coerce').astype('Int64')
            elif data_type == 'float':
                # Handle float with coercion
                variable_params[col] = pd.to_numeric(
                    variable_params[col], errors='coerce')
            elif data_type == 'datetime64' or data_type == 'date':
                # Handle datetime parsing
                variable_params[col] = try_parse_date(variable_params[col])
            elif data_type == 'str' or data_type == str:
                # Handle string with proper NA handling
                variable_params[col] = variable_params[col].fillna(
                    '').astype(str)
            else:
                # Fallback: try astype, but catch errors
                try:
                    variable_params[col] = variable_params[col].astype(
                        data_type)
                except Exception:
                    # If the specified type fails, keep as object
                    variable_params[col] = variable_params[col].astype(
                        'object')
        except Exception as e:
            print(f"Issue observed in data type {data_type} in {_parm}")
            print(f"Error as {e}")
            # Instead of failing, keep the column as object and continue
            variable_params[col] = variable_params[col].astype('object')
            continue

    # Reset index to ensure clean, sequential indices starting from 0
    variable_params = variable_params.reset_index(drop=True)

    return variable_params


def load_parameters(paramPath):
    """
    Load parameter data from Excel files in a specified path and process each sheet to convert data types.

    Parameters:
    - paramPath (str): Path to the directory containing Excel files with parameter data.

    Returns:
    - Dict: A dictionary containing processed parameter data from each sheet, where the keys are sheet names.
    """
    param = dict()

    for _wbFile in itertools.chain(paramPath.glob('[!~]*.xlsx')):
        _sheetList = load_workbook(_wbFile).sheetnames
        for _parm in _sheetList:
            df = pd.read_excel(_wbFile, sheet_name=_parm)
            try:
                param[_parm] = _deal_sheet(df, _parm)
            except Exception as e:
                print(f"Issue observed in sheet {_parm}")
                print(f"Error as {e}")
    return param


def load_configuration_file(configPath):
    """
    Load and parse a JSON configuration file for ECL run background settings.

    Parameters:
    - configPath (str): Path to the JSON configuration file.

    Returns:
    - Dict: Dictionary containing the parsed configuration settings for ECL run background.
    """
    with open(Path(configPath), 'r') as fp:
        c = json.load(fp)

    return c


if __name__ == "__main__":
    configPath = Path(
        r'C:\Users\WH947CH\Engagement\Khan Bank\03_ECL_engine\02_Development\khb_engine\run_config_file.json')
    c = load_configuration_file(configPath=configPath)

    print(c)
    # paramPath = Path(r'C:\Users\WH947CH\Engagement\Khan Bank\99_Data_Server\parameter\20240331')
    # parameters = load_parameters(paramPath)
    # print(parameters.keys())
